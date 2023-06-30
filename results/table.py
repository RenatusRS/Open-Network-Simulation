from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from parameters import SimulationParameters, Variables
from util import average_results

def tabela_analitika_simulacija(analyze_results, simulate_results):
	workbook = generate_template_table()
	
	data = pack_data(analyze_results, simulate_results)
	
	fill_table(workbook["Analitika"], data[0])
	fill_table(workbook["Simulacija x1"], data[1])
	fill_table(workbook["Odstupanja x1"], data[2])
	fill_table(workbook[f"Simulacija x{SimulationParameters.NumberOfSimulations.value}"], data[3])
	fill_table(workbook[f"Odstupanja x{SimulationParameters.NumberOfSimulations.value}"], data[4])
	
	workbook.save("results/odstupanja.xlsx")
	
	
def generate_template_table():
	workbook = Workbook()

	# Create a base template worksheet
	base_template = workbook.active
	base_template.title = "Template"
	
	orange = PatternFill(start_color="ff9900", end_color="ff9900", fill_type="solid")
	
	# Set column headers in the base template
	headers_row = [
		"K", "r",
		"Up", "Us1", "Us2", "Us3", "Uk1", "Uk2", "Uk3", "Uk4", "Uk5",
		"Jp", "Js1", "Js2", "Js3", "Jk1", "Jk2", "Jk3", "Jk4", "Jk5",
		"Xp", "Xs1", "Xs2", "Xs3", "Xk1", "Xk2", "Xk3", "Xk4", "Xk5",
		"T"
	]
	for col_num, header in enumerate(headers_row, start=1):
		column_letter = get_column_letter(col_num)
		base_template[f"{column_letter}1"] = header
		base_template[f"{column_letter}1"].font = Font(bold=True)
		base_template[f"{column_letter}1"].alignment = Alignment(horizontal="center")
		base_template[f"{column_letter}1"].fill = orange
	
	headers_column = [[K, r] for K in Variables.K.value for r in Variables.r.value]
	for row_num, header in enumerate(headers_column, start=2):
		base_template[f"A{row_num}"] = header[0]
		base_template[f"B{row_num}"] = header[1]
		base_template[f"A{row_num}"].font = Font(bold=True)
		base_template[f"B{row_num}"].font = Font(bold=True)
		base_template[f"A{row_num}"].fill = orange
		base_template[f"B{row_num}"].fill = orange
	
	# Create multiple pages using the base template
	page_names = [
		"Analitika", 
		"Simulacija x1",
		"Odstupanja x1",
		f"Simulacija x{SimulationParameters.NumberOfSimulations.value}",
		f"Odstupanja x{SimulationParameters.NumberOfSimulations.value}"
	]
	for page_name in page_names:
		worksheet = workbook.copy_worksheet(base_template)
		worksheet.title = page_name
	
	# Remove the base template worksheet
	workbook.remove(base_template)
	
	return workbook

def fill_table(worksheet, data: list):
	light_gray = PatternFill(start_color="b7b7b7", end_color="b7b7b7", fill_type="solid")
	
	for row_num, row in enumerate(data, start=2):
		for col_num, cell in enumerate(row, start=3):
			column_letter = get_column_letter(col_num)
			worksheet[f"{column_letter}{row_num}"] = cell
			worksheet[f"{column_letter}{row_num}"].fill = light_gray
			
def pack_data(analyze_results, simulate_results):
	ar = {
		K: {
			r: analyze_results[K]['parameters'][r] for r in Variables.r.value
		} for K in Variables.K.value
	}
	
	data_analitika = [[
		ar[K][r]['U']['Processor'],
		ar[K][r]['U']['System Disk 1'],
		ar[K][r]['U']['System Disk 2'],
		ar[K][r]['U']['System Disk 3'],
		ar[K][r]['U']['User Disk'] if K >= 1 else "-",
		ar[K][r]['U']['User Disk'] if K >= 2 else "-",
		ar[K][r]['U']['User Disk'] if K >= 3 else "-",
		ar[K][r]['U']['User Disk'] if K >= 4 else "-",
		ar[K][r]['U']['User Disk'] if K >= 5 else "-",
		
		ar[K][r]['J']['Processor'],
		ar[K][r]['J']['System Disk 1'],
		ar[K][r]['J']['System Disk 2'],
		ar[K][r]['J']['System Disk 3'],
		ar[K][r]['J']['User Disk'] if K >= 1 else "-",
		ar[K][r]['J']['User Disk'] if K >= 2 else "-",
		ar[K][r]['J']['User Disk'] if K >= 3 else "-",
		ar[K][r]['J']['User Disk'] if K >= 4 else "-",
		ar[K][r]['J']['User Disk'] if K >= 5 else "-",
		
		ar[K][r]['X']['Processor'],
		ar[K][r]['X']['System Disk 1'],
		ar[K][r]['X']['System Disk 2'],
		ar[K][r]['X']['System Disk 3'],
		ar[K][r]['X']['User Disk'] if K >= 1 else "-",
		ar[K][r]['X']['User Disk'] if K >= 2 else "-",
		ar[K][r]['X']['User Disk'] if K >= 3 else "-",
		ar[K][r]['X']['User Disk'] if K >= 4 else "-",
		ar[K][r]['X']['User Disk'] if K >= 5 else "-",
		
		ar[K][r]['T_OVERALL']
	] for K in Variables.K.value for r in Variables.r.value]
	
	one_result_each = {
		K: {
			r: None for r in Variables.r.value
		} for K in Variables.K.value
	}
	
	for result in simulate_results:
		one_result_each[result['K']][result['r']] = result
	
	data_simulacija_x1 = [[
		one_result_each[K][r]['utilization']['Processor'],
		one_result_each[K][r]['utilization']['System Disk 1'],
		one_result_each[K][r]['utilization']['System Disk 2'],
		one_result_each[K][r]['utilization']['System Disk 3'],
		one_result_each[K][r]['utilization']['User Disk 1'] if K >= 1 else "-",
		one_result_each[K][r]['utilization']['User Disk 2'] if K >= 2 else "-",
		one_result_each[K][r]['utilization']['User Disk 3'] if K >= 3 else "-",
		one_result_each[K][r]['utilization']['User Disk 4'] if K >= 4 else "-",
		one_result_each[K][r]['utilization']['User Disk 5'] if K >= 5 else "-",
		
		one_result_each[K][r]['processing_time']['Processor'],
		one_result_each[K][r]['processing_time']['System Disk 1'],
		one_result_each[K][r]['processing_time']['System Disk 2'],
		one_result_each[K][r]['processing_time']['System Disk 3'],
		one_result_each[K][r]['processing_time']['User Disk 1'] if K >= 1 else "-",
		one_result_each[K][r]['processing_time']['User Disk 2'] if K >= 2 else "-",
		one_result_each[K][r]['processing_time']['User Disk 3'] if K >= 3 else "-",
		one_result_each[K][r]['processing_time']['User Disk 4'] if K >= 4 else "-",
		one_result_each[K][r]['processing_time']['User Disk 5'] if K >= 5 else "-",
		
		one_result_each[K][r]['throughput']['Processor'],
		one_result_each[K][r]['throughput']['System Disk 1'],
		one_result_each[K][r]['throughput']['System Disk 2'],
		one_result_each[K][r]['throughput']['System Disk 3'],
		one_result_each[K][r]['throughput']['User Disk 1'] if K >= 1 else "-",
		one_result_each[K][r]['throughput']['User Disk 2'] if K >= 2 else "-",
		one_result_each[K][r]['throughput']['User Disk 3'] if K >= 3 else "-",
		one_result_each[K][r]['throughput']['User Disk 4'] if K >= 4 else "-",
		one_result_each[K][r]['throughput']['User Disk 5'] if K >= 5 else "-",
		
		one_result_each[K][r]['response_time']
	] for K in Variables.K.value for r in Variables.r.value]
	
	data_odstupanja_x1 = [[
		f"=ABS('Analitika'!C{row_num1 * 4 + row_num2} - 'Simulacija x1'!C{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!D{row_num1 * 4 + row_num2} - 'Simulacija x1'!D{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!E{row_num1 * 4 + row_num2} - 'Simulacija x1'!E{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!F{row_num1 * 4 + row_num2} - 'Simulacija x1'!F{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!G{row_num1 * 4 + row_num2} - 'Simulacija x1'!G{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!H{row_num1 * 4 + row_num2} - 'Simulacija x1'!H{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!I{row_num1 * 4 + row_num2} - 'Simulacija x1'!I{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!J{row_num1 * 4 + row_num2} - 'Simulacija x1'!J{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!K{row_num1 * 4 + row_num2} - 'Simulacija x1'!K{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!L{row_num1 * 4 + row_num2} - 'Simulacija x1'!L{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!M{row_num1 * 4 + row_num2} - 'Simulacija x1'!M{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!N{row_num1 * 4 + row_num2} - 'Simulacija x1'!N{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!O{row_num1 * 4 + row_num2} - 'Simulacija x1'!O{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!P{row_num1 * 4 + row_num2} - 'Simulacija x1'!P{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!Q{row_num1 * 4 + row_num2} - 'Simulacija x1'!Q{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!R{row_num1 * 4 + row_num2} - 'Simulacija x1'!R{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!S{row_num1 * 4 + row_num2} - 'Simulacija x1'!S{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!T{row_num1 * 4 + row_num2} - 'Simulacija x1'!T{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!U{row_num1 + row_num2 * 4} - 'Simulacija x1'!U{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!V{row_num1 + row_num2 * 4} - 'Simulacija x1'!V{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!W{row_num1 + row_num2 * 4} - 'Simulacija x1'!W{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!X{row_num1 + row_num2 * 4} - 'Simulacija x1'!X{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!Y{row_num1 + row_num2 * 4} - 'Simulacija x1'!Y{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!Z{row_num1 + row_num2 * 4} - 'Simulacija x1'!Z{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!AA{row_num1 + row_num2 * 4} - 'Simulacija x1'!AA{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!AB{row_num1 + row_num2 * 4} - 'Simulacija x1'!AB{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!AC{row_num1 + row_num2 * 4} - 'Simulacija x1'!AC{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!AD{row_num1 * 4 + row_num2} - 'Simulacija x1'!AD{row_num1 * 4 + row_num2})"
	] for row_num1, K in enumerate(Variables.K.value) for row_num2, r in enumerate(Variables.r.value, start=2)]
	
	sr = average_results(simulate_results)
	
	data_simulacija_xN = [[
		sr[K][r]['utilization']['Processor'],
		sr[K][r]['utilization']['System Disk 1'],
		sr[K][r]['utilization']['System Disk 2'],
		sr[K][r]['utilization']['System Disk 3'],
		sr[K][r]['utilization']['User Disk 1'] if K >= 1 else "-",
		sr[K][r]['utilization']['User Disk 2'] if K >= 2 else "-",
		sr[K][r]['utilization']['User Disk 3'] if K >= 3 else "-",
		sr[K][r]['utilization']['User Disk 4'] if K >= 4 else "-",
		sr[K][r]['utilization']['User Disk 5'] if K >= 5 else "-",
		
		sr[K][r]['processing_time']['Processor'],
		sr[K][r]['processing_time']['System Disk 1'],
		sr[K][r]['processing_time']['System Disk 2'],
		sr[K][r]['processing_time']['System Disk 3'],
		sr[K][r]['processing_time']['User Disk 1'] if K >= 1 else "-",
		sr[K][r]['processing_time']['User Disk 2'] if K >= 2 else "-",
		sr[K][r]['processing_time']['User Disk 3'] if K >= 3 else "-",
		sr[K][r]['processing_time']['User Disk 4'] if K >= 4 else "-",
		sr[K][r]['processing_time']['User Disk 5'] if K >= 5 else "-",
		
		sr[K][r]['throughput']['Processor'],
		sr[K][r]['throughput']['System Disk 1'],
		sr[K][r]['throughput']['System Disk 2'],
		sr[K][r]['throughput']['System Disk 3'],
		sr[K][r]['throughput']['User Disk 1'] if K >= 1 else "-",
		sr[K][r]['throughput']['User Disk 2'] if K >= 2 else "-",
		sr[K][r]['throughput']['User Disk 3'] if K >= 3 else "-",
		sr[K][r]['throughput']['User Disk 4'] if K >= 4 else "-",
		sr[K][r]['throughput']['User Disk 5'] if K >= 5 else "-",
		
		sr[K][r]['response_time']
	] for K in Variables.K.value for r in Variables.r.value]
	
	data_odstupanja_xN = [[
		f"=ABS('Analitika'!C{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!C{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!D{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!D{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!E{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!E{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!F{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!F{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!G{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!G{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!H{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!H{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!I{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!I{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!J{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!J{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!K{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!K{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!L{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!L{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!M{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!M{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!N{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!N{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!O{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!O{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!P{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!P{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!Q{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!Q{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!R{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!R{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!S{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!S{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!T{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!T{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!U{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!U{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!V{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!V{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!W{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!W{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!X{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!X{row_num1 * 4 + row_num2})",
		f"=ABS('Analitika'!Y{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!Y{row_num1 * 4 + row_num2})" if K >= 1 else "-",
		f"=ABS('Analitika'!Z{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!Z{row_num1 * 4 + row_num2})" if K >= 2 else "-",
		f"=ABS('Analitika'!AA{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!AA{row_num1 * 4 + row_num2})" if K >= 3 else "-",
		f"=ABS('Analitika'!AB{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!AB{row_num1 * 4 + row_num2})" if K >= 4 else "-",
		f"=ABS('Analitika'!AC{row_num1 + row_num2 * 4} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!AC{row_num1 * 4 + row_num2})" if K >= 5 else "-",
		
		f"=ABS('Analitika'!AD{row_num1 * 4 + row_num2} - 'Simulacija x{SimulationParameters.NumberOfSimulations.value}'!AD{row_num1 * 4 + row_num2})"
	] for row_num1, K in enumerate(Variables.K.value) for row_num2, r in enumerate(Variables.r.value, start=2)]
	
	return [data_analitika, data_simulacija_x1, data_odstupanja_x1, data_simulacija_xN, data_odstupanja_xN]